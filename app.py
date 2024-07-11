from flask import Flask, request, render_template, redirect, url_for, session
from openpyxl import load_workbook
import os
import firebase_config
import firebase_admin
from firebase_admin import auth
import requests

app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(24)  # Needed for session management
app.config['UPLOAD_FOLDER'] = '/tmp'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Use the firebase_ref from firebase_config module
ref = firebase_config.firebase_ref.child('all_data')
ref_statuses = firebase_config.firebase_ref.child('statuses')


def get_user_info(uid):
    user = auth.get_user(uid)
    return user

@app.route('/')
def home():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # Fetch users from Firebase to pass to the template
    ref_users = firebase_config.firebase_ref.child('users')
    users = ref_users.get()
    message = None

    if users:
        ref_accounts = firebase_config.firebase_ref.child('accounts')
        accounts = [user_data.get('account', '') for user_data in users.values()]
        accounts_string = ', '.join(accounts)

        try:
            ref_accounts.update({'accounts': accounts_string})
            message = 'Accounts updated successfully!'
        except Exception as err:
            message = f"Error updating accounts: {err}"
    else:
        message = "No users found in Firebase."

    statuses_message = None
    if users:
        statuses = [user_data.get('status', False) for user_data in users.values()]
        statuses_string = ', '.join(str(status) for status in statuses)

        try:
            ref_statuses.update({'statuses': statuses_string})
            statuses_message = 'Statuses updated successfully!'
        except Exception as err:
            statuses_message = f"Error updating statuses: {err}"
    else:
        statuses_message = "No users found in Firebase."

    return render_template('index.html', users=users, message=message, statuses_message=statuses_message)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        try:
            # Use Firebase's REST API to authenticate
            response = requests.post(f'https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={firebase_config.apiKey}',
                                     data={'email': email, 'password': password, 'returnSecureToken': True})
            response_data = response.json()
            if 'error' in response_data:
                raise Exception(response_data['error']['message'])
            
            session['user_id'] = response_data['localId']
            return redirect(url_for('home'))
        except Exception as e:
            return render_template('login.html', error=str(e))
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    return redirect(url_for('login'))

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    if 'file' not in request.files:
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)
    if file:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        # Load the Excel file
        wb = load_workbook(filename=file_path)
        sheet = wb.active

        # Read headers from the first row
        headers = [cell.value for cell in sheet[1]]

        # Read data from the Excel sheet and store it in a dictionary
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)

        columns_to_extract = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        exceldata = {}
        for column_letter in columns_to_extract:
            column_data = []
            for row in range(2, 12):  # Rows 2 to 11
                cell_value = sheet[column_letter + str(row)].value
                if cell_value is None:
                    cell_value = ""
                column_data.append(cell_value)
            exceldata[column_letter] = column_data

        final_string = ""
        for column in columns_to_extract:
            column_data = ", ".join(str(value) for value in exceldata[column])
            final_string += f"{column_data}/"

        final_string = final_string.rstrip('/')
        wb.close()

        try:
            ref.set(final_string)
            message = 'Cloud Data updated successfully!'
        except Exception as err:
            message = f"Error updating data: {err}"

        return render_template('result.html', message=message, data=final_string, headers=headers, rows=data)

@app.route('/add_user', methods=['POST'])
def add_user():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    name = request.form['name']
    account = request.form['account']

    ref_users = firebase_config.firebase_ref.child('users')
    users = ref_users.get()
    if users:
        for user_id, user_data in users.items():
            if user_data.get('name') == name or user_data.get('account') == account:
                return "exists"

    new_user = {
        'name': name,
        'account': account,
        'status': True
    }
    ref_users.push(new_user)
    return "success"

@app.route('/update_user/<user_id>', methods=['POST'])
def update_user(user_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    name = request.form['name']
    account = request.form['account']
    status = request.form.get('status') == 'on'
    ref_user = firebase_config.firebase_ref.child('users').child(user_id)
    ref_user.update({
        'name': name,
        'account': account,
        'status': status
    })
    return redirect(url_for('home'))

@app.route('/update_status/<user_id>', methods=['POST'])
def update_status(user_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    status = request.json['status']
    ref_user = firebase_config.firebase_ref.child('users').child(user_id)
    ref_user.update({'status': status})

    try:
        ref_statuses.update({'statuses': status})
        return redirect(url_for('home'))
    except Exception as e:
        return f"Error updating status: {e}"
    return '', 204

@app.route('/delete_user/<user_id>', methods=['POST'])
def delete_user(user_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    ref_user = firebase_config.firebase_ref.child('users').child(user_id)
    ref_user.delete()
    return '', 204

@app.route('/update_accounts', methods=['POST'])
def update_accounts():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    ref_users = firebase_config.firebase_ref.child('users')
    users = ref_users.get()

    if not users:
        return "No users found in Firebase."

    accounts = []
    for user_id, user_data in users.items():
        accounts.append(user_data.get('account', ''))

    accounts_string = ', '.join(accounts)

    try:
        ref.set(accounts_string)
        message = 'Accounts updated successfully!'
    except Exception as err:
        message = f"Error updating accounts: {err}"

    return redirect(url_for('result', accounts_string=accounts_string, message=message))

if __name__ == '__main__':
    app.run(debug=True, port=5555)
